"""
Data processing utilities for Severity Visualisation Dashboard.
Handles CSV parsing, feature extraction, scoring logic, and Excel export.
"""

import ast
import base64
import io

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Side → input_images dict key mapping ────────────────────────────────────

SIDE_TO_KEY = {
    "back": "BACKMERGED",
    "left": "LEFT",
    "right": "RIGHT",
    "top": "TOP",
    "bottom": "BOTTOM",
    "front": "FRONT",
    "frontblack": "FRONTBLACK",
}


def get_input_image_url(input_images_raw, side: str) -> str:
    """Return the input image URL for a given side from the input_images dict."""
    try:
        data = ast.literal_eval(str(input_images_raw))
        key = SIDE_TO_KEY.get(str(side).strip().lower(), str(side).strip().upper())
        return data.get(key, "")
    except Exception:
        return ""


# ── Severity helpers ─────────────────────────────────────────────────────────

def get_score_color(score) -> str:
    """Return a hex color for a given gt_score."""
    try:
        s = float(score)
    except (TypeError, ValueError):
        return "#94a3b8"
    if s >= 67:
        return "#ef4444"   # red
    elif s >= 34:
        return "#f59e0b"   # amber
    return "#22c55e"        # green


def get_severity_label(score) -> str:
    try:
        s = float(score)
    except (TypeError, ValueError):
        return "Unknown"
    if s >= 67:
        return "High"
    elif s >= 34:
        return "Medium"
    return "Low"


def get_severity_badge_color(score) -> str:
    """Return a Bootstrap color name for a given score."""
    try:
        s = float(score)
    except (TypeError, ValueError):
        return "secondary"
    if s >= 67:
        return "danger"
    elif s >= 34:
        return "warning"
    return "success"


# ── CSV processing ────────────────────────────────────────────────────────────

def parse_csv_content(contents: str, filename: str) -> pd.DataFrame:
    """Decode a base64-encoded dcc.Upload file content and return a DataFrame."""
    _content_type, content_string = contents.split(",", 1)
    decoded = base64.b64decode(content_string)
    df = pd.read_csv(io.StringIO(decoded.decode("utf-8")))
    return df


def enrich_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Add derived columns: input_image_url, severity, score_color."""
    df = df.copy()
    df["input_image_url"] = df.apply(
        lambda r: get_input_image_url(r.get("input_images", ""), r.get("side", "")),
        axis=1,
    )
    df["severity"] = df["gt_score"].apply(get_severity_label)
    df["score_color"] = df["gt_score"].apply(get_score_color)
    return df


def get_numeric_feature_columns(df: pd.DataFrame) -> list[str]:
    """Return numeric columns that are likely model features (exclude IDs, scores)."""
    skip = {
        "gt_score", "pre_training_score", "score_pred", "pred_cls_prob",
        "image_h", "image_w", "gt_found", "pred_cls",
    }
    return [
        c for c in df.select_dtypes(include=[np.number]).columns
        if c not in skip
    ]


def compute_correlations(df: pd.DataFrame) -> pd.Series:
    """Pearson correlation of all numeric features against gt_score."""
    feat_cols = get_numeric_feature_columns(df)
    numeric = df[feat_cols + ["gt_score"]].dropna()
    corr = numeric.corr()["gt_score"].drop("gt_score", errors="ignore")
    return corr.sort_values(key=abs, ascending=False)


# ── Excel export ──────────────────────────────────────────────────────────────

_SCORE_FILLS = {
    "High":    PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
    "Medium":  PatternFill(start_color="FFE5B4", end_color="FFE5B4", fill_type="solid"),
    "Low":     PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"),
    "Unknown": PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid"),
}

_HEADER_FILL = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_LINK_FONT   = Font(color="0563C1", underline="single")
_CENTER      = Alignment(horizontal="center", vertical="center")
_VCENTER     = Alignment(vertical="center")


def generate_excel_report(df: pd.DataFrame) -> bytes:
    """
    Build an Excel workbook with one row per record sorted by gt_score desc.
    Includes image_uuid column when available.
    """
    has_uuid = "image_uuid" in df.columns

    if has_uuid:
        headers    = ["#", "pdd_txn_id", "image_uuid", "side", "question",
                      "gt_score", "severity", "score_pred",
                      "input_image_url", "result_image_url"]
        col_widths = [5, 38, 38, 10, 35, 12, 10, 12, 65, 65]
        gt_col     = 6   # 1-indexed
        url_cols   = {9, 10}
    else:
        headers    = ["#", "pdd_txn_id", "side", "question",
                      "gt_score", "severity", "score_pred",
                      "input_image_url", "result_image_url"]
        col_widths = [5, 38, 10, 35, 12, 10, 12, 65, 65]
        gt_col     = 5
        url_cols   = {8, 9}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Severity Report"

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 30

    sorted_df = df.sort_values("gt_score", ascending=False).reset_index(drop=True)

    for rank, row in sorted_df.iterrows():
        r        = rank + 2
        gt_raw   = row.get("gt_score")
        sp_raw   = row.get("score_pred")
        severity = get_severity_label(gt_raw)

        try:    gt_val = round(float(gt_raw), 2)
        except: gt_val = ""
        try:    sp_val = round(float(sp_raw), 2)
        except: sp_val = ""

        if has_uuid:
            values = [
                rank + 1,
                str(row.get("pdd_txn_id", "")),
                str(row.get("image_uuid", "") or ""),
                str(row.get("side", "")),
                str(row.get("question", "")),
                gt_val, severity, sp_val,
                str(row.get("input_image_url", "")),
                str(row.get("result_image_url", "")),
            ]
        else:
            values = [
                rank + 1,
                str(row.get("pdd_txn_id", "")),
                str(row.get("side", "")),
                str(row.get("question", "")),
                gt_val, severity, sp_val,
                str(row.get("input_image_url", "")),
                str(row.get("result_image_url", "")),
            ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.alignment = _VCENTER
            if col_idx == gt_col:
                cell.fill = _SCORE_FILLS.get(severity, _SCORE_FILLS["Unknown"])
                cell.font = Font(bold=True)
            elif col_idx in url_cols and val:
                cell.hyperlink = val
                cell.font = _LINK_FONT

        ws.row_dimensions[r].height = 18

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── Redash UUID fetch ─────────────────────────────────────────────────────────

def fetch_uuids_from_redash(
    df: pd.DataFrame,
    redash_base_url: str,
    api_key: str,
    query_id: str,
) -> tuple[pd.DataFrame, int, int]:
    """
    Fetch image UUIDs from Redash and add an `image_uuid` column to df.

    Returns (enriched_df, redash_row_count, missing_uuid_count).
    Rows with no UUID found are kept in the df (image_uuid stays None/NaN).
    """
    import time
    import requests

    pdd_txn_ids     = df["pdd_txn_id"].dropna().unique().tolist()
    pdd_txn_ids_str = "'" + "','".join(str(x) for x in pdd_txn_ids) + "'"
    params          = {"pdd_txn_ids": pdd_txn_ids_str}

    run_url = f"{redash_base_url.rstrip('/')}/api/queries/{query_id}/results"
    headers = {
        "Authorization": f"Key {api_key}",
        "Content-Type": "application/json",
    }

    run_resp = requests.post(run_url, headers=headers, json={"parameters": params}, timeout=30)
    run_resp.raise_for_status()
    resp_json = run_resp.json()

    # Resolve result — may be cached or need job polling
    if "query_result" in resp_json:
        qr       = resp_json["query_result"]
        data_rows = qr["data"]["rows"]
        columns   = qr["data"]["columns"]
    else:
        job = resp_json.get("job")
        if not job:
            raise RuntimeError(f"Unexpected Redash response: {resp_json}")

        job_id     = job["id"]
        status_url = f"{redash_base_url.rstrip('/')}/api/jobs/{job_id}"
        max_wait   = 120
        waited     = 0
        query_result_id = None

        while waited < max_wait:
            s     = requests.get(status_url, headers=headers, timeout=15).json()
            state = s["job"]["status"]
            if state == 3:      # success
                query_result_id = s["job"]["query_result_id"]
                break
            elif state == 4:    # failed
                raise RuntimeError(f"Redash job failed: {s}")
            time.sleep(2)
            waited += 2

        if query_result_id is None:
            raise RuntimeError("Redash query timed out after 120 s")

        results_url = (
            f"{redash_base_url.rstrip('/')}/api/query_results/"
            f"{query_result_id}.json?api_key={api_key}"
        )
        results  = requests.get(results_url, timeout=30).json()
        qr       = results["query_result"]
        data_rows = qr["data"]["rows"]
        columns   = qr["data"]["columns"]

    redash_df    = pd.DataFrame(data_rows)
    column_order = [c["name"] for c in columns]
    redash_df    = redash_df[[c for c in column_order if c in redash_df.columns]]

    # Build (pdd_txn_id, side) → uuid mapping
    uuid_mapping: dict[tuple, str] = {}
    for _, rrow in redash_df.iterrows():
        txn = rrow.get("pdd_txn_id")
        if pd.isna(txn):
            continue
        for side in ["left", "right", "top", "bottom", "back", "front"]:
            col = f"{side}_uuid"
            if col in redash_df.columns:
                val = rrow.get(col)
                if pd.notna(val) and str(val).strip():
                    uuid_mapping[(str(txn), side)] = str(val).strip()

    df = df.copy()
    df["image_uuid"] = df.apply(
        lambda r: uuid_mapping.get(
            (str(r["pdd_txn_id"]), str(r["side"]).lower().strip())
        )
        if pd.notna(r.get("pdd_txn_id")) and pd.notna(r.get("side"))
        else None,
        axis=1,
    )

    missing = int(df["image_uuid"].isna().sum())
    return df, len(redash_df), missing
